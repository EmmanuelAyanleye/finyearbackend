from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import CustomUser
from .forms import LecturerRegistrationForm, StudentUpdateForm
from django.contrib.auth.decorators import user_passes_test
from .forms import StudentForm
from django.contrib.sessions.models import Session
from django.utils import timezone
from django.urls import reverse

from .models import Lecturer
from .forms import CourseForm

from django.shortcuts import redirect
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.contrib.auth.hashers import make_password
import json
from .models import Course, Attendance
from fingerprint.models import FingerprintData  # Import the FingerprintData model
import pytz


from .forms import StudentSelfUpdateForm

from .models import CustomUser, Student, Department, AcademicSession
from .forms import LecturerProfileUpdateForm




def group_required(group_name):
    def decorator(view_func):
        def _wrapped_view(request, *args, **kwargs):
            if not request.user.groups.filter(name=group_name).exists():
                return redirect('home')  # Redirect to home page if not in group
            return view_func(request, *args, **kwargs)
        return _wrapped_view
    return decorator

# views.py
def is_admin(user):
    return user.is_authenticated and user.role == 'admin'

def is_lecturer(user):
    return user.is_authenticated and user.role == 'lecturer'

def is_student(user):
    return user.is_authenticated and user.role == 'student'




# 🔹 Login View
def login_view(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')

        user = authenticate(request, email=email, password=password)

        if user is not None:
            # Get the user's active session (if any)
            active_session = Session.objects.filter(
                expire_date__gte=timezone.now()
            ).filter(
                session_key=request.session.session_key
            )

            if active_session.exists():
                messages.error(request, 'You are already logged in on another tab!')
                return redirect('login')

            # Create a new session
            request.session.save()

            login(request, user)

            # Redirect user based on role
            if user.role == 'admin':
                return redirect('admin_dashboard')
            elif user.role == 'student':
                return redirect('student_panel')
            elif user.role == 'lecturer':
                return redirect('lecturer_panel')

        else:
            messages.error(request, 'Invalid credentials')

    return render(request, 'home/login.html')



def logout_view(request):
    logout(request)
    request.session.pop('active_user', None)
    return redirect('login')














def index(request):
    return render(request, 'home/index.html')  # Render the homepage template



from django.db.models import Q

def get_courses(request):
    department_id = request.GET.get('department')
    level = request.GET.get('level')

    if department_id and level:
        # Get both department-specific and general courses
        courses = Course.objects.filter(
            (Q(departments__id=department_id) | Q(is_general=True)) & Q(level=level)
        ).distinct().values('id', 'course_code', 'course_title')

        # Debug print to check what courses are being returned
        print(f"Found {len(courses)} courses for department {department_id} and level {level}")
        for course in courses:
            print(f"Course: {course['course_code']} - {course['course_title']}")

        course_list = [
            {
                'id': course['id'],
                'course_title': f"{course['course_code']} - {course['course_title']}"
            }
            for course in courses
        ]
        
        return JsonResponse(course_list, safe=False)

    return JsonResponse([], safe=False)



from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.utils.timezone import localtime
from .models import Student, Course, Attendance

from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from django.utils.timezone import localtime
from .models import Student, Course, Attendance, Department

def mark(request):
    if request.method == "POST":
        print("Received POST data:", request.POST)  # Debugging line

        department = request.POST.get("department")
        level = request.POST.get("level")
        course_id = request.POST.get("course")

        if not department or not level or not course_id:
            messages.error(request, "All fields are required")
            return redirect("mark")

        # Store selected course and session in the session
        request.session["selected_course_id"] = course_id
        request.session["selected_session"] = level  # Assuming you want to store the level as session data

        print("Redirecting to verify...")  # Debugging line before the redirect
        return redirect("verify")

    departments = Department.objects.all()
    return render(request, "home/mark.html", {"departments": departments})





from datetime import datetime

from django.utils import timezone

from datetime import datetime

def verify(request):
    course_id = request.session.get("selected_course_id")
    session = request.session.get("selected_session")

    if not course_id or not session:
        return redirect("mark")

    course = get_object_or_404(Course, id=course_id)

    # Get current day of the week
    current_day = timezone.localtime(timezone.now()).strftime('%A')  # Returns day name like 'Monday'
    
    # Convert the times to 'time' objects if they are not already
    course_start_time = course.attendance_start_time.time() if isinstance(course.attendance_start_time, datetime) else course.attendance_start_time
    course_end_time = course.attendance_end_time.time() if isinstance(course.attendance_end_time, datetime) else course.attendance_end_time

    # Get the current time (ensure it is in the same timezone)
    now = timezone.localtime(timezone.now()).time()

    # Debug: print times and days
    print("Current day:", current_day)
    print("Course day:", course.attendance_day)
    print("Current time:", now)
    print("Course start time:", course_start_time)
    print("Course end time:", course_end_time)

    # Check if current day matches course day
    if current_day.lower() != course.attendance_day.lower():
        messages.error(request, f"This class is scheduled for {course.attendance_day}s only")

    # Check if current time is within the allowed attendance timeframe
    if not (course_start_time <= now <= course_end_time):
        messages.error(request, "Not yet time for the class")
        print("Not within the allowed time range.")

    # Only return the template if both day and time are valid
    if current_day.lower() != course.attendance_day.lower() or not (course_start_time <= now <= course_end_time):
        return render(request, "home/verify.html", {"course": course, "session": session})

    return render(request, "home/verify.html", {"course": course, "session": session})


def get_dashboard_stats(request):
    today = timezone.now().date()
    
    # Get total number of students
    total_students = Student.objects.count()
    
    # Get total attendance count
    total_attendance = Attendance.objects.count()
    
    # Get today's attendance count
    todays_attendance = Attendance.objects.filter(date=today).count()
    
    # Calculate percentage changes (you may want to adjust the comparison period)
    prev_month_attendance = Attendance.objects.filter(
        date__month=(today.month - 1 if today.month > 1 else 12)
    ).count()
    
    if prev_month_attendance > 0:
        attendance_change = ((total_attendance - prev_month_attendance) / prev_month_attendance) * 100
    else:
        attendance_change = 0
        
    # Today's attendance change compared to yesterday
    yesterday_attendance = Attendance.objects.filter(
        date=(today - timezone.timedelta(days=1))
    ).count()
    
    if yesterday_attendance > 0:
        today_change = ((todays_attendance - yesterday_attendance) / yesterday_attendance) * 100
    else:
        today_change = 0

    return JsonResponse({
        'total_students': total_students,
        'total_attendance': total_attendance,
        'todays_attendance': todays_attendance,
        'attendance_change': round(attendance_change, 1),
        'today_change': round(today_change, 1)
    })

from channels.generic.websocket import WebsocketConsumer
import json
from asgiref.sync import async_to_sync

def fingerprint(request):
    course_id = request.session.get("selected_course_id")
    session = request.session.get("selected_session")
    if not course_id or not session:
        return redirect("mark")
    course = Course.objects.get(id=course_id)
    
    # Get the latest verified student details from session if available
    student_details = request.session.get('verified_student_details', None)
    
    context = {
        "course": course, 
        "session": session,
        "student_details": student_details if student_details else None
    }
    
    # Clear the session data after using it
    if student_details:
        del request.session['verified_student_details']
    
    return render(request, "home/fingerprint.html", context)


# home/views.py
from django.shortcuts import render

@login_required
@user_passes_test(is_admin)
def course(request):
    if request.method == 'POST':
        form = CourseForm(request.POST)
        if form.is_valid():
            try:
                # Check if course code already exists
                course_code = form.cleaned_data['course_code']
                if Course.objects.filter(course_code=course_code).exists():
                    messages.error(request, f'Course code "{course_code}" already exists')
                    return render(request, 'home/course.html', {
                        'form': form,
                        'departments': Department.objects.all()
                    })

                # Create course without saving to DB
                course = form.save(commit=False)
                
                # Set is_general flag
                course.is_general = request.POST.get('is_general_course') == 'true'
                
                # Save the course to get an ID
                course.save()

                # Handle department assignments
                if course.is_general:
                    # Add all departments if it's a general course
                    departments = Department.objects.all()
                    course.departments.set(departments)
                else:
                    dept_ids = request.POST.get('selected_departments', '').split(',')
                    if dept_ids and dept_ids[0]:  # Check if there are selected departments
                        departments = Department.objects.filter(id__in=dept_ids)
                        course.departments.set(departments)

                messages.success(request, 'Course added successfully')
                return redirect('course')
            except Exception as e:
                messages.error(request, f'There is an existing course code: {str(e)}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = CourseForm()
    
    context = {
        'form': form,
        'departments': Department.objects.all()
    }
    return render(request, 'home/course.html', context)


@user_passes_test(is_admin)
def report(request):
    # Get all unique values for filters
    departments = Department.objects.all()
    courses = Course.objects.all()
    sessions = Attendance.objects.values_list('session__name', flat=True).distinct()
    semesters = Attendance.objects.values_list('semester__name', flat=True).distinct()
    
    context = {
        'departments': departments,
        'courses': courses,
        'sessions': sessions,
        'semesters': semesters,
    }
    return render(request, 'home/report.html', context)

def get_attendance_data(request):
    # Get filter parameters
    search_query = request.GET.get('search', '')
    department = request.GET.get('department', '')
    level = request.GET.get('level', '')
    course = request.GET.get('course', '')
    semester = request.GET.get('semester', '')
    session = request.GET.get('session', '')

    # Base queryset
    queryset = Attendance.objects.select_related(
        'student', 'student__department', 'course'
    ).order_by('-date', '-timestamp')

    # Apply filters
    if search_query:
        queryset = queryset.filter(
            Q(student__full_name__icontains=search_query) |
            Q(student__matric_number__icontains=search_query)
        )
    if department:
        queryset = queryset.filter(student__department__name=department)
    if level:
        queryset = queryset.filter(student__level=level)
    if course:
        queryset = queryset.filter(course__course_code=course)
    if semester:
        queryset = queryset.filter(semester__name=semester)
    if session:
        queryset = queryset.filter(session__name=session)

    # Prepare data for response
    attendance_data = []
    for idx, attendance in enumerate(queryset, 1):
        # Convert timestamp to local time
        local_timestamp = timezone.localtime(attendance.timestamp)
        
        attendance_data.append({
            'id': idx,
            'student_name': attendance.student.full_name,
            'department': attendance.student.department.name,
            'level': attendance.student.level,
            'course': f"{attendance.course.course_code} - {attendance.course.course_title}",
            'semester': attendance.semester.name,
            'session': attendance.session.name,
            'date': local_timestamp.strftime('%Y-%m-%d'),
            'timestamp': local_timestamp.strftime('%I:%M %p'),  # Use local time
            'status': attendance.status
        })

    return JsonResponse({'data': attendance_data})

def export_excel(request):
    # Get filtered data
    search_query = request.GET.get('search', '')
    department = request.GET.get('department', '')
    level = request.GET.get('level', '')
    course = request.GET.get('course', '')
    semester = request.GET.get('semester', '')
    session = request.GET.get('session', '')

    # Create workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Define headers
    headers = ['#', 'Student Name', 'Department', 'Level', 'Course', 
              'Semester', 'Session', 'Date', 'Time', 'Status']
    
    # Style headers
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = openpyxl.styles.Font(color='FFFFFF', bold=True)

    # Get data using the same filter logic from get_attendance_data
    queryset = Attendance.objects.select_related(
        'student', 'student__department', 'course'
    ).order_by('-date', '-timestamp')

    # Apply filters
    if search_query:
        queryset = queryset.filter(
            Q(student__full_name__icontains=search_query) |
            Q(student__matric_number__icontains=search_query)
        )
    if department:
        queryset = queryset.filter(student__department__name=department)
    if level:
        queryset = queryset.filter(student__level=level)
    if course:
        queryset = queryset.filter(course__course_code=course)
    if semester:
        queryset = queryset.filter(semester__name=semester)
    if session:
        queryset = queryset.filter(session__name=session)

    # Write data
    for idx, attendance in enumerate(queryset, 1):
        row = [
            idx,
            attendance.student.full_name,
            attendance.student.department.name,
            attendance.student.level,
            f"{attendance.course.course_code} - {attendance.course.course_title}",
            attendance.semester.name,
            attendance.session.name,
            attendance.date.strftime('%Y-%m-%d'),
            attendance.timestamp.strftime('%I:%M %p'),
            attendance.status
        ]
        ws.append(row)

        # Color status cells
        status_cell = ws.cell(row=idx+1, column=10)
        if attendance.status == 'present':
            status_cell.fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
        elif attendance.status == 'absent':
            status_cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=attendance_report.xlsx'
    
    wb.save(response)
    return response



@login_required(login_url='login')
@user_passes_test(is_admin, login_url='index')
def admin_dashboard(request):
    context = {
        # other context variables...
    }
    return render(request, 'home/admin_dashboard.html', context)


@user_passes_test(is_admin)
def admin_summary(request):
    # Get data from database
    departments = Department.objects.all()
    courses = Course.objects.all()
    semesters = Semester.objects.all()
    sessions = Session.objects.all()
    
    context = {
        'departments': departments,
        'courses': courses,
        'semesters': semesters,
        'sessions': sessions,
    }
    return render(request, 'home/admin-summary.html', context)

def get_attendance_summary(request):
    # Get filter parameters
    department = request.GET.get('department', '')
    level = request.GET.get('level', '')
    course = request.GET.get('course', '')
    semester = request.GET.get('semester', '')
    session = request.GET.get('session', '')

    # Base queryset
    queryset = Attendance.objects.values(
        'student__id',
        'student__full_name',
        'student__matric_number'
    ).annotate(
        total_attendance=Count('id')
    ).order_by('-total_attendance')

    # Apply filters
    if department:
        queryset = queryset.filter(student__department__name=department)
    if level:
        queryset = queryset.filter(student__level=level)
    if course:
        queryset = queryset.filter(course__course_code=course)
    if semester:
        queryset = queryset.filter(semester__name=semester)
    if session:
        queryset = queryset.filter(session__name=session)

    # Prepare summary data
    summary_data = []
    for idx, record in enumerate(queryset, 1):
        summary_data.append({
            'id': idx,
            'student_name': record['student__full_name'],
            'matric_number': record['student__matric_number'],
            'total_attendance': record['total_attendance']
        })

    # Get total number of students
    total_students = len(summary_data)

    return JsonResponse({
        'data': summary_data,
        'total_students': total_students
    })

def export_summary(request):
    # Get filter parameters and create Excel file similar to export_excel
    # But with summary format instead
    # ... implementation similar to export_excel but with summary data ...
    pass

# def add_lecturer(request):
#     return render(request, 'home/add-lecturer.html')


import subprocess
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt



@user_passes_test(is_admin)
def add_student(request):
    departments = Department.objects.all()
    sessions = AcademicSession.objects.all()

    if request.method == "POST":
        full_name = request.POST.get("full_name")
        matric_number = request.POST.get("matric_number")
        email = request.POST.get("email")
        password = request.POST.get("password")
        session_id = request.POST.get("session")
        department_id = request.POST.get("department")
        level = request.POST.get("level")
        profile_picture = request.FILES.get("profile_picture")
        fingerprint_template = request.POST.get("fingerprint1")
        gender = request.POST.get('gender')

        if not fingerprint_template:
            messages.error(request, "Fingerprint data is required!")
            return redirect("add_student")

        # Check for duplicates
        if CustomUser.objects.filter(email=email).exists():
            messages.error(request, "Email already exists!")
            return redirect("add_student")

        if Student.objects.filter(matric_number=matric_number).exists():
            messages.error(request, "Matric number already exists!")
            return redirect("add_student")

        # Create user and student record
        user = CustomUser.objects.create_user(email=email, password=password, role="student")
        session = AcademicSession.objects.get(id=session_id)
        department = Department.objects.get(id=department_id)

        student = Student.objects.create(
            user=user,
            full_name=full_name,
            matric_number=matric_number,
            session=session,
            department=department,
            level=level,
            profile_picture=profile_picture,
            gender=gender
        )

        FingerprintData.objects.create(student=student, template=fingerprint_template)

        messages.success(request, "Student added successfully!")
        return redirect("add_student")

    return render(request, "home/add-student.html", {"departments": departments, "sessions": sessions})


@user_passes_test(is_admin)
def modify_lecturer(request):
    search_query = request.GET.get('search', '')  # Get search query from URL parameters
    lecturers = Lecturer.objects.all()

    if search_query:
        lecturers = lecturers.filter(
            full_name__icontains=search_query) | lecturers.filter(
            user__email__icontains=search_query) | lecturers.filter(
            lecturer_id__icontains=search_query
        )

    return render(request, 'home/modify-lecturer.html', {'lecturers': lecturers})

@user_passes_test(is_admin)
def delete_lecturer(request, lecturer_id):
    if request.method == "POST":
        try:
            lecturer = Lecturer.objects.get(id=lecturer_id)
            
            # Check if lecturer has assigned courses
            assigned_courses = Course.objects.filter(lecturer=lecturer)
            
            if assigned_courses.exists():
                # Get list of course codes
                course_codes = [course.course_code for course in assigned_courses]
                
                return JsonResponse({
                    "success": False,
                    "message": f"Cannot delete lecturer. Please reassign these courses first: {', '.join(course_codes)}"
                })
            
            # If no courses, proceed with deletion
            user = lecturer.user
            lecturer.delete()
            user.delete()
            
            return JsonResponse({"success": True})
            
        except Lecturer.DoesNotExist:
            return JsonResponse({
                "success": False, 
                "message": "Lecturer not found"
            })
        except Exception as e:
            return JsonResponse({
                "success": False, 
                "message": str(e)
            })
    
    return JsonResponse({
        "success": False, 
        "message": "Invalid request method"
    })


@user_passes_test(is_admin)
def modify_lecturer_page(request, lecturer_id):
    lecturer = get_object_or_404(Lecturer, lecturer_id=lecturer_id)

    if request.method == "POST":
        full_name = request.POST.get('full_name')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        password = request.POST.get('password')
        
        # Update lecturer details
        lecturer.full_name = full_name
        lecturer.phone = phone
        lecturer.user.email = email

        if password:
            # Set password if provided
            lecturer.user.set_password(password)

        # Handle profile picture upload
        profile_picture = request.FILES.get('profile_picture')
        if profile_picture:
            lecturer.profile_picture = profile_picture

        # Save the changes
        lecturer.user.save()
        lecturer.save()

        # Show success message
        messages.success(request, "Lecturer details updated successfully!")
        return redirect('modify_lecturer_page', lecturer_id=lecturer_id)

    return render(request, 'home/modify-lecturer-page.html', {'lecturer': lecturer})


@user_passes_test(is_admin)
def modify_course(request):
    search_query = request.GET.get('search', '')  # Get search query from URL parameters
    courses = Course.objects.all()

    if search_query:
        courses = courses.filter(
            course_title__icontains=search_query) | courses.filter(
            course_code__icontains=search_query) | courses.filter(
            lecturer__full_name__icontains=search_query
        )

    return render(request, 'home/modify-course.html', {'courses': courses})

@user_passes_test(is_admin)
def delete_course(request, course_id):
    if request.method == "POST":
        try:
            course = Course.objects.get(id=course_id)
            course_code = course.course_code

            if course.delete():
                return JsonResponse({"success": True})
            else:
                return JsonResponse({"success": False, "message": "Error deleting course"})
        except Course.DoesNotExist:
            return JsonResponse({"success": False, "message": "Course not found"})
        except Exception as e:
            return JsonResponse({"success": False, "message": str(e)})
    return JsonResponse({"success": False, "message": "Invalid request"})


@login_required
@user_passes_test(is_admin)
def delete_session(request, pk):
    if request.method == "POST":
        try:
            session = AcademicSession.objects.get(pk=pk)
            session.delete()
            return JsonResponse({"success": True})
        except AcademicSession.DoesNotExist:
            return JsonResponse({"success": False, "message": "Session not found"})
        except Exception as e:
            return JsonResponse({"success": False, "message": str(e)})
    return JsonResponse({"success": False, "message": "Invalid request"})

@login_required
@user_passes_test(is_admin)
def delete_semester(request, pk):
    if request.method == "POST":
        try:
            semester = Semester.objects.get(pk=pk)
            semester.delete()
            return JsonResponse({"success": True})
        except Semester.DoesNotExist:
            return JsonResponse({"success": False, "message": "Semester not found"})
        except Exception as e:
            return JsonResponse({"success": False, "message": str(e)})
    return JsonResponse({"success": False, "message": "Invalid request"})

@login_required
@user_passes_test(is_admin)
def delete_department(request, pk):
    if request.method == "POST":
        try:
            department = Department.objects.get(pk=pk)
            department.delete()
            return JsonResponse({"success": True})
        except Department.DoesNotExist:
            return JsonResponse({"success": False, "message": "Department not found"})
        except Exception as e:
            return JsonResponse({"success": False, "message": str(e)})
    return JsonResponse({"success": False, "message": "Invalid request"})



from datetime import datetime, time
from django.utils import timezone
from .models import Course, Student, Attendance
import pytz

@user_passes_test(is_admin)
@csrf_exempt
def remove_student_course(request, student_id, course_id):
    student = get_object_or_404(Student, id=student_id)
    course = get_object_or_404(Course, id=course_id)

    if request.method == "POST":
        # Remove attendance records linked to this course
        Attendance.objects.filter(student=student, course=course).delete()

        # Remove the course from the student
        student.courses.remove(course)

        return JsonResponse({"success": True})

    return JsonResponse({"success": False})



@user_passes_test(is_admin)
def view_student_courses(request, student_id):
    student = get_object_or_404(Student, id=student_id)

    # Get automatic courses (department + level)
    department_courses = Course.objects.filter(
        departments=student.department,
        level=student.level,
        is_general=False
    )

    # Get general courses for the student's level
    general_courses = Course.objects.filter(
        is_general=True,
        level=student.level
    )

    # Get courses manually assigned by the admin
    assigned_courses = student.courses.all()

    # Merge all courses while avoiding duplicates
    all_courses = list(set(department_courses) | set(general_courses) | set(assigned_courses))

    # Fetch all courses that are NOT yet assigned to the student
    available_courses = Course.objects.exclude(id__in=[course.id for course in all_courses])

    context = {
        "student": student,
        "courses": all_courses,  # All courses the student is currently offering
        "all_courses": available_courses,  # Courses available for assignment
    }

    return render(request, "home/view_student_courses.html", context)



@csrf_exempt
def assign_courses_to_student(request):
    if request.method == "POST":
        student_id = request.POST.get('student_id')
        course_ids = request.POST.getlist('courses')

        student = get_object_or_404(Student, id=student_id)

        if course_ids:
            courses = Course.objects.filter(id__in=course_ids)
            student.courses.add(*courses)
            return JsonResponse({'success': True})

    return JsonResponse({'success': False})


@login_required
def student_courses(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    
    # Get courses specific to the student's department and level (compulsory courses)
    department_courses = list(Course.objects.filter(
        departments=student.department,
        level=student.level,
        is_general=False
    ))

    # Get general courses available for the student's level
    general_courses = list(Course.objects.filter(
        is_general=True,
        level=student.level
    ))

    # Get courses manually assigned by the admin
    assigned_courses = list(student.courses.all())  # Courses already assigned

    # Merge all three lists and remove duplicates
    all_courses = list({course.id: course for course in department_courses + general_courses + assigned_courses}.values())

    # Sort by course code
    all_courses.sort(key=lambda x: x.course_code)

    context = {
        'student': student,
        'courses': all_courses,  # Display all relevant courses
        'total_courses': len(all_courses)
    }

    return render(request, 'home/student_courses.html', context)




# home/tasks.py (Create this file)
from celery import shared_task
from django.utils import timezone
from .models import Course, Student, Attendance
import pytz
from datetime import datetime

@shared_task
def mark_absent_students():
    """
    Marks students as 'absent' for courses where they didn't mark attendance.
    This task should be scheduled to run periodically (e.g., daily).
    """
    lagos_tz = pytz.timezone('Africa/Lagos')
    current_time = timezone.now().astimezone(lagos_tz)
    current_day = current_time.strftime('%A')
    today_date = current_time.date()

    # Get all courses scheduled for today
    today_courses = Course.objects.filter(attendance_day__iexact=current_day)

    for course in today_courses:
        # Check if the course's end time has passed
        if current_time.time() > course.attendance_end_time.time():
            # Get all students enrolled in this course
            if course.is_general:
                enrolled_students = Student.objects.filter(level=course.level)
            else:
                enrolled_students = Student.objects.filter(
                    department__in=course.departments.all(),
                    level=course.level
                ).distinct()

            # Get students who have already marked attendance for this course today
            marked_students = Attendance.objects.filter(
                course=course,
                date=today_date
            ).values_list('student_id', flat=True)

            # Mark absent for students who haven't marked attendance
            for student in enrolled_students:
                if student.id not in marked_students:
                    Attendance.objects.create(
                        student=student,
                        course=course,
                        status='absent',
                        date=today_date,
                        semester=course.semester,
                        session=student.session,
                        timestamp=current_time
                    )

import io
import os
from datetime import datetime, time, timedelta
from django.http import FileResponse, HttpResponse
from django.shortcuts import get_object_or_404
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from .models import Course, Department  # Import Course and Department models
from django.db.models import Q

def generate_timetable_pdf(request):
    """
    Generates a timetable PDF based on the courses in the database.
    """
    try:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=50,
            leftMargin=50,
            topMargin=50,
            bottomMargin=50
        )
        elements = []
        styles = getSampleStyleSheet()

        # --- School Header ---
        logo_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "static", "Images", "Espam logo.png")
        try:
            elements.append(Image(logo_path, width=70, height=70))
        except:
            pass  # Continue if logo is missing

        header_style = ParagraphStyle('Header', parent=styles['Title'], fontSize=16, alignment=1, spaceAfter=6)
        elements.append(Paragraph("ESPAM FORMATION UNIVERSITY", header_style))
        subtext_style = ParagraphStyle('Subtext', parent=styles['Normal'], fontSize=10, alignment=1)
        elements.append(Paragraph("Sacouba Anavie Campus, Porto-Novo, Republic of Benin", subtext_style))
        elements.append(Paragraph("Phone: +22946436904, +2348035637035, +22956885802", subtext_style))
        elements.append(Paragraph("Email: espamformationunicampus2@gmail.com", subtext_style))
        elements.append(Spacer(1, 10))

        # --- Timetable Title ---
        report_title_style = ParagraphStyle('ReportTitle', parent=styles['Heading2'], fontSize=14, alignment=1, spaceAfter=10)
        elements.append(Paragraph("COURSE TIMETABLE", report_title_style))
        elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%d %B, %Y')}", subtext_style))
        elements.append(Spacer(1, 20))

        # --- Timetable Data ---
        # Dynamically create time slots (only start times)
        courses = Course.objects.all().order_by('attendance_day', 'attendance_start_time')
        all_start_times = sorted(list(set([course.attendance_start_time for course in courses])))

        # Use only start times for time slots
        time_slots = all_start_times
        time_slot_labels = [t.strftime("%I:%M %p") for t in time_slots]

        # Define days of the week
        days_of_week = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

        # Create the timetable data structure
        timetable_data = [["DAY/TIME"] + time_slot_labels]  # Header row
        for day in days_of_week:
            day_row = [day]
            for time_slot in time_slots:
                day_row.append("")  # Initialize empty cells
            timetable_data.append(day_row)

        # Populate the timetable
        for course in courses:
            day = course.attendance_day
            start_time = course.attendance_start_time
            end_time = course.attendance_end_time

            # Find the correct row and column for the course
            day_index = days_of_week.index(day) + 1  # +1 to account for the header row
            try:
                start_time_index = time_slots.index(start_time) + 1  # +1 to account for the "DAY/TIME" column
            except ValueError:
                print(f"Error: Start time {start_time} not found in time_slots.")
                continue  # Skip this course if start time is not found

            # Add course details to the cells
            course_info = f"{course.course_code}"  # Only course code

            # Add course to the start slot only
            if timetable_data[day_index][start_time_index] == "":
                timetable_data[day_index][start_time_index] = course_info
            else:
                timetable_data[day_index][start_time_index] += f"\n{course_info}"

        # --- Create and Style the Table ---
        table = Table(timetable_data, colWidths=[80] + [60] * len(time_slots))
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),  # Header row background
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),  # Header row text color
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Center align all cells
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Header row font
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),  # Header row padding
            ('BACKGROUND', (0, 1), (0, -1), colors.lightgrey), # Day column background
            ('GRID', (0, 0), (-1, -1), 1, colors.black),  # Add grid lines
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(table)

        # --- Build the PDF ---
        doc.build(elements)
        buffer.seek(0)

        return FileResponse(
            buffer, as_attachment=True, filename=f"timetable_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        )

    except Exception as e:
        return HttpResponse(f"Error generating PDF: {str(e)}", status=500)


@user_passes_test(is_admin)
def modify_course_page(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    lecturers = Lecturer.objects.all()
    departments = Department.objects.all()
    semesters = Semester.objects.all()

    if request.method == "POST":
        try:
            # Update basic course information
            course.course_code = request.POST.get('courseCode')
            course.course_title = request.POST.get('courseTitle')
            course.semester_id = request.POST.get('semester')
            course.level = request.POST.get('level')
            course.lecturer_id = request.POST.get('lecturers')
            course.attendance_day = request.POST.get('attendanceDay')
            course.attendance_start_time = request.POST.get('attendanceStartTime')
            course.attendance_end_time = request.POST.get('attendanceEndTime')
            
            # Handle department assignments
            is_general = request.POST.get('is_general_course') == 'true'
            course.is_general = is_general
            
            if is_general:
                # Add all departments if it's a general course
                course.departments.set(departments)
            else:
                # Add selected departments
                dept_ids = request.POST.get('selected_departments', '').split(',')
                if dept_ids and dept_ids[0]:
                    selected_departments = Department.objects.filter(id__in=dept_ids)
                    course.departments.set(selected_departments)
                else:
                    messages.error(request, 'Please select at least one department')
                    return render(request, 'home/modify-course-page.html', {
                        'course': course,
                        'lecturers': lecturers,
                        'departments': departments,
                        'semesters': semesters
                    })

            course.save()
            messages.success(request, 'Course updated successfully!')
            
            # Stay on the same page after success
            return render(request, 'home/modify-course-page.html', {
                'course': course,
                'lecturers': lecturers,
                'departments': departments,
                'semesters': semesters
            })
            
        except Exception as e:
            messages.error(request, f'Error updating course: {str(e)}')
            return render(request, 'home/modify-course-page.html', {
                'course': course,
                'lecturers': lecturers,
                'departments': departments,
                'semesters': semesters
            })
            
    context = {
        'course': course,
        'lecturers': lecturers,
        'departments': departments,
        'semesters': semesters
    }
    return render(request, 'home/modify-course-page.html', context)


@user_passes_test(is_admin)
def modify_student(request):
    students = Student.objects.select_related('user', 'department', 'session').all()
    return render(request, 'home/modify-student.html', {'students': students})



@csrf_exempt
@user_passes_test(is_admin)
def delete_student(request, student_id):
    try:
        student = Student.objects.get(id=student_id)
        student.user.delete()  # This will also delete the related student record
        return JsonResponse({"success": True})
    except Student.DoesNotExist:
        return JsonResponse({"success": False, "error": "Student not found"})


@user_passes_test(is_admin)
def modify_student_page(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    sessions = AcademicSession.objects.all().order_by('-name')  # Get all sessions
    
    if request.method == 'POST':
        form = StudentUpdateForm(request.POST, request.FILES, instance=student)
        if form.is_valid():
            try:
                # Get form data
                student.full_name = form.cleaned_data['full_name']
                student.matric_number = form.cleaned_data['matric_number']
                student.gender = form.cleaned_data['gender']
                student.level = form.cleaned_data['level']
                student.session = form.cleaned_data['session']
                student.department = form.cleaned_data['department']
                
                # Update user email if changed
                if student.user.email != form.cleaned_data['email']:
                    student.user.email = form.cleaned_data['email']
                    student.user.save()
                
                # Update password if provided
                if form.cleaned_data.get('password'):
                    student.user.set_password(form.cleaned_data['password'])
                    student.user.save()
                
                student.save()
                messages.success(request, 'Student details updated successfully!')
                return redirect('modify_student_page' , student_id=student_id)
            except Exception as e:
                messages.error(request, f'Error updating student details: {str(e)}')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = StudentUpdateForm(instance=student, initial={
            'email': student.user.email,
            'session': student.session,
        })

    context = {
        'student': student,
        'form': form,
        'sessions': sessions,
    }
    return render(request, 'home/modify-student-page.html', context)


from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db.models import Count
from .models import Lecturer, Course, Student, Attendance


from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db.models import Count
from home.models import Course, Lecturer, Student, Attendance

import logging

logger = logging.getLogger(__name__)
@login_required
def lecturer_panel(request):
    try:
        lecturer = get_object_or_404(Lecturer, user=request.user)
        
        # Fetch courses assigned to the lecturer
        courses = Course.objects.filter(lecturer=lecturer)
        course_ids = courses.values_list('id', flat=True)  # Get course IDs
        
        # Get distinct students from the assigned courses
        students = Student.objects.filter(courses__in=courses).distinct()
        
        # Count attendance records for courses under the lecturer
        attendance = Attendance.objects.filter(course_id__in=course_ids)

        context = {
            'lecturer': lecturer,
            'total_courses': courses.count(),
            'total_students': students.count(),
            'total_attendance': attendance.count(),
        }

        print("DEBUG: Lecturer Panel Context:", context)  # Debugging print statement

        return render(request, 'home/lecturer-panel.html', context)

    except Exception as e:
        print(f"Error in lecturer_panel: {str(e)}")
        return render(request, 'home/lecturer-panel.html', {
            'lecturer': None,
            'total_courses': 0,
            'total_students': 0,
            'total_attendance': 0
        })



@login_required
def get_lecturer_stats(request):
    try:
        lecturer = get_object_or_404(Lecturer, user=request.user)
        
        # Fetch data
        courses = Course.objects.filter(lecturer=lecturer)
        students = Student.objects.filter(courses__in=courses).distinct()
        attendance = Attendance.objects.filter(course__in=courses)

        # Response Data
        data = {
            'total_courses': courses.count(),
            'total_students': students.count(),
            'total_attendance': attendance.count(),
        }
        
        return JsonResponse(data)

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)



@login_required
@user_passes_test(is_lecturer)
def lecturer_report(request):
    lecturer = get_object_or_404(Lecturer, user=request.user)
    departments = Department.objects.all()
    sessions = AcademicSession.objects.all()
    semesters = Semester.objects.all()
    # Only get courses assigned to this lecturer
    courses = Course.objects.filter(lecturer=lecturer)
    
    context = {
        'departments': departments,
        'sessions': sessions,
        'semesters': semesters,
        'courses': courses,
        'lecturer': lecturer
    }
    return render(request, 'home/lecturer-report.html', context)

@login_required
@user_passes_test(is_lecturer)
def get_lecturer_attendance_data(request):
    try:
        lecturer = get_object_or_404(Lecturer, user=request.user)
        
        # Get filter parameters
        search = request.GET.get('search', '')
        department = request.GET.get('department')
        session = request.GET.get('session')
        semester = request.GET.get('semester')
        course = request.GET.get('course')
        
        # Start with attendance for lecturer's courses
        queryset = Attendance.objects.filter(course__lecturer=lecturer)
        
        # Apply filters
        if search:
            queryset = queryset.filter(student__full_name__icontains=search)
        if department:
            queryset = queryset.filter(student__department__name=department)
        if session:
            queryset = queryset.filter(session__name=session)
        if semester:
            queryset = queryset.filter(semester__name=semester)
        if course:
            queryset = queryset.filter(course__course_code=course)

        # Prepare data for response
        attendance_data = []
        for idx, attendance in enumerate(queryset, 1):
            local_time = timezone.localtime(attendance.timestamp)
            attendance_data.append({
                'id': idx,
                'student_name': attendance.student.full_name,
                'department': attendance.student.department.name,
                'level': attendance.student.level,
                'course': f"{attendance.course.course_code} - {attendance.course.course_title}",
                'semester': attendance.semester.name,
                'session': attendance.session.name,
                'date': local_time.strftime('%Y-%m-%d'),
                'time': local_time.strftime('%I:%M %p'),
                'status': attendance.status
            })

        return JsonResponse({'data': attendance_data})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@user_passes_test(is_lecturer)
def export_lecturer_attendance(request):
    try:
        lecturer = get_object_or_404(Lecturer, user=request.user)
        
        # Get filtered data
        search = request.GET.get('search', '')
        department = request.GET.get('department')
        session = request.GET.get('session')
        semester = request.GET.get('semester')
        course = request.GET.get('course')
        
        # Get filtered queryset
        queryset = Attendance.objects.filter(course__lecturer=lecturer)
        
        if department:
            queryset = queryset.filter(student__department__name=department)
        if session:
            queryset = queryset.filter(session__name=session)
        if semester:
            queryset = queryset.filter(semester__name=semester)
        if course:
            queryset = queryset.filter(course__course_code=course)

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        # Add headers with styling
        headers = ['#', 'Student Name', 'Department', 'Level', 'Course', 
                  'Semester', 'Session', 'Date', 'Time', 'Status']
        
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = Font(color='FFFFFF', bold=True)

        # Add data
        for idx, attendance in enumerate(queryset, 1):
            local_time = timezone.localtime(attendance.timestamp)
            row = [
                idx,
                attendance.student.full_name,
                attendance.student.department.name,
                attendance.student.level,
                f"{attendance.course.course_code} - {attendance.course.course_title}",
                attendance.semester.name,
                attendance.session.name,
                local_time.strftime('%Y-%m-%d'),
                local_time.strftime('%I:%M %p'),
                attendance.status
            ]
            ws.append(row)

            # Color status cells
            status_cell = ws.cell(row=idx+1, column=10)
            if attendance.status == 'present':
                status_cell.fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
            elif attendance.status == 'absent':
                status_cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # Create response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=attendance_report.xlsx'
        
        wb.save(response)
        return response
        
    except Exception as e:
        return HttpResponse(f"Error exporting data: {str(e)}", status=500) 


from django.http import JsonResponse
from django.db.models import Count
from django.contrib.auth.decorators import login_required, user_passes_test
from .models import Lecturer, Course, Student, Attendance, AcademicSession, Semester

@login_required
@user_passes_test(is_lecturer)
def lecturer_summary(request):
    try:
        lecturer = get_object_or_404(Lecturer, user=request.user)
        
        # Get available sessions, semesters and courses for this lecturer
        sessions = AcademicSession.objects.all().order_by('-name')
        semesters = Semester.objects.all()
        courses = Course.objects.filter(lecturer=lecturer)
        
        context = {
            'lecturer': lecturer,
            'sessions': sessions,
            'semesters': semesters, 
            'courses': courses,
            'default_course': courses.first() # Get first course as default
        }
        
        return render(request, "home/lecturer-summary.html", context)
    except Exception as e:
        print(f"Error: {str(e)}")
        return render(request, "home/lecturer-summary.html", {})

@login_required
@user_passes_test(is_lecturer)
def get_lecturer_summary_data(request):
    try:
        lecturer = get_object_or_404(Lecturer, user=request.user)
        course_id = request.GET.get('course')
        session_id = request.GET.get('session')
        semester_id = request.GET.get('semester')
        
        # Base query for courses
        courses = Course.objects.filter(lecturer=lecturer)
        if course_id:
            courses = courses.filter(id=course_id)
        
        # Get students with attendance records
        students_with_attendance = Student.objects.filter(
            attendance__course__in=courses
        ).distinct()
        
        summary_data = []
        
        for student in students_with_attendance:
            attendance_query = Attendance.objects.filter(
                student=student,
                course__in=courses,
                status='present'
            )
            
            if session_id:
                attendance_query = attendance_query.filter(session_id=session_id)
            if semester_id:
                attendance_query = attendance_query.filter(semester_id=semester_id)
            
            total_attendance = attendance_query.count()
            
            summary_data.append({
                'student_name': student.full_name,
                'matric_number': student.matric_number,
                'total_attendance': total_attendance,
            })
        
        # Sort by student name
        summary_data.sort(key=lambda x: x['student_name'])
        
        course_name = None
        if course_id:
            course = courses.first()
            course_name = f"{course.course_code} - {course.course_title}"
        
        return JsonResponse({
            'data': summary_data,
            'total_students': len(summary_data),
            'course_name': course_name or 'All Courses'
        })
        
    except Exception as e:
        print(f"Error in get_lecturer_summary_data: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)
    



@user_passes_test(is_lecturer)
def manage_class(request):
    # Get the lecturer object
    lecturer = Lecturer.objects.get(user=request.user)

    # Fetch courses assigned to the logged-in lecturer
    courses = Course.objects.filter(lecturer=lecturer)

    return render(request, 'home/manage-class.html', {'courses': courses}) 


@user_passes_test(is_lecturer)
def modify_class(request, course_id):
    course_allocation = get_object_or_404(Course, id=course_id)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            new_day = data.get('day')
            new_start_time = data.get('start_time')
            new_end_time = data.get('end_time')

            # Validate data
            if not new_day or not new_start_time or not new_end_time:
                return JsonResponse({'error': 'All fields are required'}, status=400)

            # Update course allocation
            course_allocation.attendance_day = new_day
            course_allocation.attendance_start_time = new_start_time
            course_allocation.attendance_end_time = new_end_time
            course_allocation.save()

            return JsonResponse({'message': 'Class updated successfully'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return render(request, 'home/modify-class.html', {
        'course': course_allocation,  # ✅ Send course details to the template
    })

from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError

User = get_user_model()
@user_passes_test(is_lecturer)
def modify_profile(request):
    try:
        lecturer = Lecturer.objects.get(user=request.user)
    except Lecturer.DoesNotExist:
        lecturer = None

    if request.method == 'POST':
        form = LecturerProfileUpdateForm(request.POST, request.FILES, instance=lecturer)
        
        if form.is_valid():
            try:
                form.save(user=request.user)  # Pass the user argument
                messages.success(request, "Profile updated successfully!")
                # Stay on the same page instead of redirecting
            except ValidationError as e:
                form.add_error('email', e)  # Add error to the form instead of redirecting
                messages.error(request, "This email is already in use by another account.")

        else:
            messages.error(request, "There was an error updating your profile.")
    
    else:
        form = LecturerProfileUpdateForm(instance=lecturer)

    return render(request, "home/modify-profile.html", {"form": form, "lecturer": lecturer})


@user_passes_test(is_lecturer)
def profile(request):
    try:
        lecturer = Lecturer.objects.get(user=request.user)  # Get lecturer details linked to the logged-in user
    except Lecturer.DoesNotExist:
        lecturer = None  # If no lecturer profile exists

    return render(request, "home/profile.html", {"lecturer": lecturer})





@login_required
@user_passes_test(is_student)
def student_summary(request):
    try:
        student = get_object_or_404(Student, user=request.user)
        
        # Get available sessions, semesters and courses for filters
        sessions = AcademicSession.objects.all().order_by('-name')
        semesters = Semester.objects.all()
        courses = Course.objects.filter(student=student
        ).distinct().order_by('course_code')

        
        context = {
            'student': student,
            'sessions': sessions,
            'semesters': semesters,
            'courses': courses
        }
        
        return render(request, 'home/student-summary.html', context)
    except Exception as e:
        print(f"Error in student_summary: {str(e)}")
        return render(request, 'home/student-summary.html', {})


@login_required
@user_passes_test(is_student)
def get_student_summary_data(request):
    try:
        student = get_object_or_404(Student, user=request.user)
        print(f"Getting summary for student: {student.full_name}")
        
        # Get filter parameters
        session_id = request.GET.get('session')
        semester_id = request.GET.get('semester')
        course_id = request.GET.get('course')
        
        # Base query for attendance records
        attendance_query = Attendance.objects.filter(student=student)
        
        # Apply filters
        if session_id:
            attendance_query = attendance_query.filter(session_id=session_id)
        if semester_id:
            attendance_query = attendance_query.filter(semester_id=semester_id)
        if course_id:
            attendance_query = attendance_query.filter(course_id=course_id)
            
        # Get unique courses with attendance
        courses = Course.objects.filter(
            attendances__in=attendance_query
        ).distinct()
        
        summary_data = []
        selected_course = None
        
        for course in courses:
            course_attendance = attendance_query.filter(course=course).count()
            
            if course_id and str(course.id) == str(course_id):
                selected_course = f"{course.course_code} - {course.course_title}"
            
            summary_data.append({
                'course': f"{course.course_code} - {course.course_title}",
                'total_classes': course_attendance
            })
        
        # Sort by course code
        summary_data.sort(key=lambda x: x['course'])
        
        response_data = {
            'data': summary_data,
            'total_courses': len(summary_data),
            'selected_course': selected_course
        }
        
        print(f"Found {len(summary_data)} courses with attendance")
        return JsonResponse(response_data)
        
    except Exception as e:
        print(f"Error in get_student_summary_data: {str(e)}")
        return JsonResponse({
            'error': str(e),
            'message': 'Failed to fetch summary data'
        }, status=500)
     



@login_required
@user_passes_test(is_student)
def student_report(request):
    try:
        student = get_object_or_404(Student, user=request.user)
        
        # Get courses this student has attendance records for
        courses = Course.objects.filter(
            attendance__student=student
        ).distinct().order_by('course_code')
        
        sessions = AcademicSession.objects.all().order_by('-name')
        semesters = Semester.objects.all()
        
        context = {
            'student': student,
            'courses': courses,
            'sessions': sessions,
            'semesters': semesters,
        }
        
        return render(request, 'home/student-report.html', context)
    except Exception as e:
        print(f"Error in student_report: {str(e)}")
        return render(request, 'home/student-report.html', {})

@login_required
@user_passes_test(is_student)
def get_student_attendance_data(request):
    try:
        # Get the logged in student
        student = get_object_or_404(Student, user=request.user)
        print(f"Fetching attendance for student: {student.full_name}")
        
        # Get filter parameters
        session_id = request.GET.get('session')
        semester_id = request.GET.get('semester')
        course_id = request.GET.get('course')
        search = request.GET.get('search', '').strip()
        
        # Base query for student's attendance
        query = Attendance.objects.filter(student=student)
        print(f"Base query count: {query.count()}")
        
        # Apply filters if provided
        if session_id:
            query = query.filter(session_id=session_id)
        if semester_id:
            query = query.filter(semester_id=semester_id)
        if course_id:
            query = query.filter(course_id=course_id)
        
        # Apply search if provided
        if search:
            from django.db.models import Q
            query = query.filter(
                Q(course__course_code__icontains=search) |
                Q(course__course_title__icontains=search) |
                Q(semester__name__icontains=search) |
                Q(session__name__icontains=search) |
                Q(date__icontains=search)
            )
        
        # Get related data in one query
        attendance_records = query.select_related(
            'course', 'semester', 'session'
        ).order_by('-date', '-timestamp')
        
        print(f"Final query count: {attendance_records.count()}")
        
        # Format data for response
        data = []
        for record in attendance_records:
            local_time = timezone.localtime(record.timestamp)
            data.append({
                'course': f"{record.course.course_code} - {record.course.course_title}",
                'semester': record.semester.name,
                'session': record.session.name,
                'date': record.date.strftime('%d-%m-%Y'),
                'time': local_time.strftime('%I:%M %p'),
                'status': record.status.title()
            })
        
        response_data = {
            'data': data,
            'total_records': len(data)
        }
        print(f"Sending {len(data)} records")
        return JsonResponse(response_data)
        
    except Exception as e:
        print(f"Error in get_student_attendance_data: {str(e)}")
        return JsonResponse({
            'error': str(e),
            'message': 'Failed to fetch attendance records'
        }, status=500)


@login_required
@user_passes_test(is_student)
def export_student_attendance_excel(request):
    try:
        student = get_object_or_404(Student, user=request.user)
        
        # Get filter parameters
        session_id = request.GET.get('session')
        semester_id = request.GET.get('semester')
        course_id = request.GET.get('course')
        
        # Base query
        attendance_records = Attendance.objects.filter(
            student=student
        ).select_related(
            'course', 'semester', 'session'
        ).order_by('-date', '-timestamp')
        
        # Apply filters
        if session_id:
            attendance_records = attendance_records.filter(session_id=session_id)
        if semester_id:
            attendance_records = attendance_records.filter(semester_id=semester_id)
        if course_id:
            attendance_records = attendance_records.filter(course_id=course_id)

        # Create workbook and sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        # Write headers with styling
        headers = ['#', 'Course', 'Semester', 'Session', 'Date', 'Time', 'Status']
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = Font(color='FFFFFF', bold=True)

        # Write data
        for idx, record in enumerate(attendance_records, 1):
            local_time = timezone.localtime(record.timestamp)
            row = [
                idx,
                f"{record.course.course_code} - {record.course.course_title}",
                record.semester.name,
                record.session.name,
                record.date.strftime('%d-%m-%Y'),
                local_time.strftime('%I:%M %p'),
                record.status.title()
            ]
            ws.append(row)

            # Color status cells
            status_cell = ws.cell(row=idx+1, column=7)  # Status is in column 7
            if record.status.lower() == 'present':
                status_cell.fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
            elif record.status.lower() == 'absent':
                status_cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # Create response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=attendance_report_{student.matric_number}.xlsx'
        
        wb.save(response)
        return response
        
    except Exception as e:
        print(f"Error generating Excel: {str(e)}")
        return HttpResponse('Error generating Excel file', status=500)

from django.contrib.auth import login

@user_passes_test(is_student)
def student_modify(request):
    student = request.user.student

    if request.method == "POST":
        form = StudentSelfUpdateForm(request.POST, request.FILES, instance=student)
        if form.is_valid():
            # Check if the new email already exists
            new_email = form.cleaned_data['email']
            if request.user.email != new_email and CustomUser.objects.filter(email=new_email).exists():
                messages.error(request, "Email already exists. Please use another email.")
            else:
                # Update user email
                request.user.email = new_email
                request.user.save()

                # Update password only if provided
                new_password = form.cleaned_data['password']
                if new_password:
                    request.user.set_password(new_password)
                    request.user.save()

                    # Re-authenticate the user after password change
                    login(request, request.user)

                # Save profile picture if changed
                form.save()
                messages.success(request, "Profile updated successfully!")
                return redirect('student_modify')

    else:
        form = StudentSelfUpdateForm(instance=student, initial={'email': request.user.email})

    return render(request, 'home/student-modify.html', {'form': form, 'student': student})

@user_passes_test(is_student)
def student_profile(request):
    student = get_object_or_404(Student, user=request.user)
    return render(request, 'home/student-profile.html', {"student": student})


@login_required(login_url='login')
@user_passes_test(is_student, login_url='index') 
def student_panel(request):
    student = request.user.student  # Get the logged-in student
    context = {
        'student': student,
    }
    return render(request, 'home/student-panel.html')


@login_required
def get_student_stats(request):
    try:
        student = get_object_or_404(Student, user=request.user)

        # Get courses specific to the student's department and level (compulsory courses)
        department_courses = Course.objects.filter(
            departments=student.department,
            level=student.level,
            is_general=False
        )

        # Get general courses available for the student's level
        general_courses = Course.objects.filter(
            is_general=True,
            level=student.level
        )

        # Get courses manually assigned by the admin (specially assigned)
        assigned_courses = student.courses.all()  # Courses already manually assigned

        # Merge all three sets of courses and remove duplicates
        all_courses = list({course.id: course for course in department_courses | general_courses | assigned_courses}.values())

        # Get the total attendance records the student has
        total_attendance = Attendance.objects.filter(student=student).count()

        # Return the data
        data = {
            'total_courses': len(all_courses),  # Total of all courses assigned (automatically and manually)
            'total_attendance': total_attendance,
        }

        return JsonResponse(data)

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)














from .forms import LecturerRegistrationForm


@user_passes_test(is_admin)
def add_lecturer(request):
    if request.method == 'POST':
        form = LecturerRegistrationForm(request.POST, request.FILES)
        
        if form.is_valid():
            email = form.cleaned_data['email']
            lecturer_id = form.cleaned_data['lecturer_id']  # Grab the lecturer_id
            
            # Check if the email already exists
            if CustomUser.objects.filter(email=email).exists():
                form.add_error('email', 'A lecturer with this email already exists.')  # Add error to the email field
                messages.error(request, 'A lecturer with this email already exists.')  # General error message
            # Check if the lecturer_id already exists
            elif Lecturer.objects.filter(lecturer_id=lecturer_id).exists():
                form.add_error('lecturer_id', 'Lecturer ID already exists!')  # Add error to the lecturer_id field
                messages.error(request, 'Lecturer ID already exists!')  # General error message
            else:
                # If the email and lecturer_id are unique, proceed to save
                user = CustomUser.objects.create_user(
                    email=email,
                    password=form.cleaned_data['password'],
                    role='lecturer'
                )
                lecturer = form.save(commit=False)
                lecturer.user = user
                lecturer.save()
                messages.success(request, "Lecturer added successfully!")
                
                # Reset the form after success
                form = LecturerRegistrationForm()  # Clear form after successful submission
        else:
            # If the form is not valid, show the generic error message
            messages.error(request, "Error adding lecturer. Please check the form.")

    else:
        form = LecturerRegistrationForm()

    return render(request, 'home/add-lecturer.html', {'form': form})











from .models import AcademicSession, Semester, Department
from .forms import AcademicSessionForm, SemesterForm, DepartmentForm
@login_required
@user_passes_test(is_admin)
def settings(request):
    if request.method == "POST":
        if "add_session" in request.POST:
            session_name = request.POST.get("session_name")
            try:
                if AcademicSession.objects.filter(name=session_name).exists():
                    messages.error(request, "This session already exists")
                else:
                    AcademicSession.objects.create(name=session_name)
                    messages.success(request, "Session added successfully")
            except Exception as e:
                messages.error(request, str(e))
                
        elif "add_semester" in request.POST:
            semester_name = request.POST.get("semester_name")
            try:
                if Semester.objects.filter(name=semester_name).exists():
                    messages.error(request, "This semester already exists")
                else:
                    Semester.objects.create(name=semester_name)
                    messages.success(request, "Semester added successfully")
            except Exception as e:
                messages.error(request, str(e))
                
        elif "add_department" in request.POST:
            department_name = request.POST.get("department_name")
            try:
                if Department.objects.filter(name=department_name).exists():
                    messages.error(request, "This department already exists")
                else:
                    Department.objects.create(name=department_name)
                    messages.success(request, "Department added successfully")
            except Exception as e:
                messages.error(request, str(e))

    context = {
        'sessions': AcademicSession.objects.all().order_by('-name'),
        'semesters': Semester.objects.all().order_by('name'),
        'departments': Department.objects.all().order_by('name'),
    }
    return render(request, 'home/settings.html', context)


from django.http import JsonResponse, HttpResponse
from openpyxl import Workbook
from django.db.models import Q, Count
import openpyxl
from openpyxl.styles import PatternFill, Font
from datetime import datetime
from .models import Course, Student, Attendance #Added to the existing import
from django.utils import timezone #Added to the existing import

@csrf_exempt
@require_http_methods(["POST"])
def mark_attendance(request):
    try:
        data = json.loads(request.body)
        matric_number = data.get('matric_number')
        course_id = data.get('course_id')
        current_session = data.get('session')

        # Get required objects
        student = Student.objects.get(matric_number=matric_number)
        course = Course.objects.get(id=course_id)

        # Check if student is enrolled in the course
        
        # Get courses specific to the student's department and level (compulsory courses)
        department_courses = list(Course.objects.filter(
            departments=student.department,
            level=student.level,
            is_general=False
        ))

        # Get general courses available for the student's level
        general_courses = list(Course.objects.filter(
            is_general=True,
            level=student.level
        ))

        # Get courses manually assigned by the admin
        assigned_courses = list(student.courses.all())  # Courses already assigned

        # Merge all three lists and remove duplicates
        all_courses = list({course.id: course for course in department_courses + general_courses + assigned_courses}.values())
        
        if course not in all_courses:
            return JsonResponse({
                "status": "error",
                "message": "You are not enrolled in this course!"
            })

        # Get the student's current academic session
        academic_session = student.session

        # Use local time for creating the attendance record
        local_tz = pytz.timezone('Africa/Lagos')
        current_time = timezone.localtime(timezone.now())

        # Check if attendance already exists
        existing_attendance = Attendance.objects.filter(
            student=student,
            course=course,
            date=timezone.now().date()
        ).exists()

        if existing_attendance:
            return JsonResponse({
                "status": "error",
                "message": "Attendance already marked for today"
            })

        # Create attendance record
        attendance = Attendance.objects.create(
            student=student,
            course=course,
            session=academic_session,
            semester=course.semester,
            status='present',
            date=current_time.date(),
            timestamp=current_time 
        )

        # Clear session data after successful attendance
        if 'verified_student' in request.session:
            del request.session['verified_student']
        if 'selected_course_id' in request.session:
            del request.session['selected_course_id']
        if 'selected_session' in request.session:
            del request.session['selected_session']

        return JsonResponse({
            "status": "success",
            "message": "Attendance marked successfully"
        })

    except Student.DoesNotExist:
        return JsonResponse({
            "status": "error",
            "message": "Student not found"
        })
    except Course.DoesNotExist:
        return JsonResponse({
            "status": "error",
            "message": "Course not found"
        })
    except Exception as e:
        return JsonResponse({
            "status": "error",
            "message": str(e)
        })
    
    
from django.http import JsonResponse, HttpResponse
from openpyxl import Workbook
from django.db.models import Q, Count
import openpyxl
from openpyxl.styles import PatternFill, Font
from datetime import datetime

def report(request):
    # Get all unique values for filters
    departments = Department.objects.all()
    courses = Course.objects.all()
    sessions = Attendance.objects.values_list('session__name', flat=True).distinct()
    semesters = Attendance.objects.values_list('semester__name', flat=True).distinct()
    
    context = {
        'departments': departments,
        'courses': courses,
        'sessions': sessions,
        'semesters': semesters,
    }
    return render(request, 'home/report.html', context)

def get_attendance_data(request):
    # Get filter parameters
    search_query = request.GET.get('search', '')
    department = request.GET.get('department', '')
    level = request.GET.get('level', '')
    course = request.GET.get('course', '')
    semester = request.GET.get('semester', '')
    session = request.GET.get('session', '')

    # Base queryset
    queryset = Attendance.objects.select_related(
        'student', 'student__department', 'course'
    ).order_by('-date', '-timestamp')

    # Apply filters
    if search_query:
        queryset = queryset.filter(
            Q(student__full_name__icontains=search_query) |
            Q(student__matric_number__icontains=search_query)
        )
    if department:
        queryset = queryset.filter(student__department__name=department)
    if level:
        queryset = queryset.filter(student__level=level)
    if course:
        queryset = queryset.filter(course__course_code=course)
    if semester:
        queryset = queryset.filter(semester__name=semester)
    if session:
        queryset = queryset.filter(session__name=session)

    # Prepare data for response
    attendance_data = []
    for idx, attendance in enumerate(queryset, 1):
        local_time = timezone.localtime(attendance.timestamp)
        attendance_data.append({
            'id': idx,
            'student_name': attendance.student.full_name,
            'department': attendance.student.department.name,
            'level': attendance.student.level,
            'course': attendance.course.course_code,
            # Code with both Course title and course code
            # 'course': f"{attendance.course.course_code} - {attendance.course.course_title}",
            'semester': attendance.semester.name,
            'session': attendance.session.name,
            'date': attendance.date.strftime('%Y-%m-%d'),
            'timestamp': local_time.strftime('%I:%M %p'),
            'status': attendance.status
        })

    return JsonResponse({'data': attendance_data})

def export_excel(request):
    # Get filtered data
    search_query = request.GET.get('search', '')
    department = request.GET.get('department', '')
    level = request.GET.get('level', '')
    course = request.GET.get('course', '')
    semester = request.GET.get('semester', '')
    session = request.GET.get('session', '')

    # Create workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Define headers
    headers = ['#', 'Student Name', 'Department', 'Level', 'Course', 
              'Semester', 'Session', 'Date', 'Time', 'Status']
    
    # Style headers
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = openpyxl.styles.Font(color='FFFFFF', bold=True)

    # Get data using the same filter logic from get_attendance_data
    queryset = Attendance.objects.select_related(
        'student', 'student__department', 'course'
    ).order_by('-date', '-timestamp')

    # Apply filters
    if search_query:
        queryset = queryset.filter(
            Q(student__full_name__icontains=search_query) |
            Q(student__matric_number__icontains=search_query)
        )
    if department:
        queryset = queryset.filter(student__department__name=department)
    if level:
        queryset = queryset.filter(student__level=level)
    if course:
        queryset = queryset.filter(course__course_code=course)
    if semester:
        queryset = queryset.filter(semester__name=semester)
    if session:
        queryset = queryset.filter(session__name=session)

    # Write data
    for idx, attendance in enumerate(queryset, 1):
        local_time = timezone.localtime(attendance.timestamp)
        row = [
            idx,
            attendance.student.full_name,
            attendance.student.department.name,
            attendance.student.level,
            f"{attendance.course.course_code} - {attendance.course.course_title}",
            attendance.semester.name,
            attendance.session.name,
            attendance.date.strftime('%Y-%m-%d'),
            local_time.strftime('%I:%M %p'),
            attendance.status
        ]
        ws.append(row)

        # Color status cells
        status_cell = ws.cell(row=idx+1, column=10)
        if attendance.status == 'present':
            status_cell.fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
        elif attendance.status == 'absent':
            status_cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=attendance_report.xlsx'
    
    wb.save(response)
    return response

from django.http import FileResponse
import io
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle




from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from datetime import datetime

import io
import os
from datetime import datetime
from django.http import FileResponse
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

def export_summary_pdf(request):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=50,
        leftMargin=50,
        topMargin=50,
        bottomMargin=50
    )
    elements = []
    styles = getSampleStyleSheet()
    
    # School Header with Logo
    logo_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "static", "Images", "Espam logo.png")

    try:
        elements.append(Image(logo_path, width=80, height=80))
    except:
        pass  # If the logo is not found, continue without it
    
    header_style = ParagraphStyle(
        'Header', parent=styles['Title'], fontSize=16, alignment=1, spaceAfter=6
    )
    elements.append(Paragraph("ESPAM FORMATION UNIVERSITY", header_style))
    
    subtext_style = ParagraphStyle(
        'Subtext', parent=styles['Normal'], fontSize=10, alignment=1
    )
    elements.append(Paragraph("Address: Sacouba Anavie Campus, Porto-Novo, Republic of Benin.", subtext_style))
    elements.append(Paragraph("Phone: +22946436904, +2348035637035, +22956885802", subtext_style))
    elements.append(Paragraph("Email: espamformationunicampus2@gmail.com", subtext_style))
    elements.append(Spacer(1, 10))
    
    # Report Title
    report_title_style = ParagraphStyle(
        'ReportTitle', parent=styles['Heading2'], fontSize=14, alignment=1, spaceAfter=10
    )
    elements.append(Paragraph("ATTENDANCE SUMMARY REPORT", report_title_style))
    elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%d %B, %Y')}", subtext_style))
    elements.append(Spacer(1, 10))
    
    # Report Filters
    department = request.GET.get('department', 'All Departments')
    level = request.GET.get('level', 'All Levels')
    course = request.GET.get('course', 'All Courses')
    semester = request.GET.get('semester', 'All Semesters')
    session = request.GET.get('session', 'All Sessions')
    
    filter_style = ParagraphStyle('FilterStyle', parent=styles['Normal'], fontSize=10, spaceAfter=5)
    filters = [
        f"Department: {department}",
        f"Level: {level}",
        f"Course: {course}",
        f"Semester: {semester}",
        f"Session: {session}"
    ]
    
    for info in filters:
        elements.append(Paragraph(info, filter_style))
    elements.append(Spacer(1, 10))
    
    # Query Attendance Data
    queryset = Attendance.objects.values(
        'student__full_name', 'student__matric_number', 'student__department__name', 'student__level'
    ).annotate(
        total_attendance=Count('id')
    ).order_by('-total_attendance')
    
    # Apply filters
    if department != 'All Departments':
        queryset = queryset.filter(student__department=department)
    if level != 'All Levels':
        queryset = queryset.filter(student__level=level)
    if course != 'All Courses':
        queryset = queryset.filter(course__course_code=course)
    if semester != 'All Semesters':
        queryset = queryset.filter(semester__name=semester)
    if session != 'All Sessions':
        queryset = queryset.filter(session__name=session)
    
    # Create Table Data
    data = [['#', 'Student Name', 'Matric No', 'Department', 'Level', 'Total Attendance']]
    for idx, record in enumerate(queryset, 1):
        data.append([
            str(idx), record['student__full_name'], record['student__matric_number'],
            record['student__department__name'], record['student__level'], str(record['total_attendance'])
        ])
    
    # Table Styling
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    
    return FileResponse(
        buffer, as_attachment=True, filename=f'attendance_summary_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf'
    )



# @login_required
# @user_passes_test(is_lecturer)
import io
import os
from datetime import datetime
from django.http import FileResponse, HttpResponse
from django.shortcuts import get_object_or_404
from django.db.models import Count
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from .models import Attendance, Lecturer, Student, Course

def export_lecturer_summary_pdf(request):
    try:
        lecturer = get_object_or_404(Lecturer, user=request.user)
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=50,
            leftMargin=50,
            topMargin=50,
            bottomMargin=50
        )
        elements = []
        styles = getSampleStyleSheet()

        # Load School Logo
        logo_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "static", "Images", "Espam logo.png")
        try:
            elements.append(Image(logo_path, width=70, height=70))
        except:
            pass  # Continue if logo is missing

        # School Header
        header_style = ParagraphStyle('Header', parent=styles['Title'], fontSize=16, alignment=1, spaceAfter=6)
        elements.append(Paragraph("ESPAM FORMATION UNIVERSITY", header_style))
        subtext_style = ParagraphStyle('Subtext', parent=styles['Normal'], fontSize=10, alignment=1)
        elements.append(Paragraph("Sacouba Anavie Campus, Porto-Novo, Republic of Benin", subtext_style))
        elements.append(Paragraph("Phone: +22946436904, +2348035637035, +22956885802", subtext_style))
        elements.append(Paragraph("Email: espamformationunicampus2@gmail.com", subtext_style))
        elements.append(Spacer(1, 10))

        # Report Title
        report_title_style = ParagraphStyle('ReportTitle', parent=styles['Heading2'], fontSize=14, alignment=1, spaceAfter=10)
        elements.append(Paragraph("ATTENDANCE SUMMARY REPORT", report_title_style))
        elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%d %B, %Y')}", subtext_style))
        elements.append(Spacer(1, 10))

        # Course Details
        course_id = request.GET.get('course')
        semester_id = request.GET.get('semester')
        session_id = request.GET.get('session')

        courses = Course.objects.filter(lecturer=lecturer)
        if course_id:
            courses = courses.filter(id=course_id)
        course_name = courses.first().course_code if courses.exists() else "All Courses"

        filter_style = ParagraphStyle('FilterStyle', parent=styles['Normal'], fontSize=10, spaceAfter=5)
        elements.append(Paragraph(f"<b>Course:</b> {course_name}", filter_style))
        elements.append(Paragraph(f"<b>Semester:</b> All Semesters", filter_style))
        elements.append(Paragraph(f"<b>Session:</b> All Sessions", filter_style))
        elements.append(Spacer(1, 10))

        # Query Students with Attendance
        students_with_attendance = Student.objects.filter(attendance__course__in=courses).distinct()
        summary_data = []

        for student in students_with_attendance:
            attendance_query = Attendance.objects.filter(student=student, course__in=courses, status="present")
            if session_id:
                attendance_query = attendance_query.filter(session_id=session_id)
            if semester_id:
                attendance_query = attendance_query.filter(semester_id=semester_id)
            total_attendance = attendance_query.count()
            summary_data.append([
                student.full_name,
                student.matric_number,
                student.department.name,
                student.level,
                total_attendance
            ])

        summary_data.sort(key=lambda x: x[0])  # Sort by student name

        # Create Table Data
        table_data = [["#", "Student Name", "Matric No", "Department", "Level", "Total Attendance"]]
        for idx, data in enumerate(summary_data, start=1):
            table_data.append([str(idx), data[0], data[1], data[2], data[3], str(data[4])])

        # Table Styling
        table = Table(table_data, colWidths=[30, 150, 100, 120, 50, 80])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(table)
        doc.build(elements)

        buffer.seek(0)
        return FileResponse(buffer, as_attachment=True, filename="attendance_summary.pdf")

    except Exception as e:
        return HttpResponse(f"Error generating PDF: {str(e)}", status=500)
    



@login_required
@user_passes_test(is_student)
def export_student_summary_pdf(request):
    try:
        student = get_object_or_404(Student, user=request.user)
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=50,
            leftMargin=50,
            topMargin=50,
            bottomMargin=50
        )
        elements = []
        styles = getSampleStyleSheet()

        # Load School Logo
        logo_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "static", "Images", "Espam logo.png")
        try:
            elements.append(Image(logo_path, width=70, height=70))
        except:
            pass  # Continue if logo is missing

        # School Header
        header_style = ParagraphStyle('Header', parent=styles['Title'], fontSize=16, alignment=1, spaceAfter=6)
        elements.append(Paragraph("ESPAM FORMATION UNIVERSITY", header_style))
        subtext_style = ParagraphStyle('Subtext', parent=styles['Normal'], fontSize=10, alignment=1)
        elements.append(Paragraph("Sacouba Anavie Campus, Porto-Novo, Republic of Benin", subtext_style))
        elements.append(Paragraph("Phone: +22946436904, +2348035637035, +22956885802", subtext_style))
        elements.append(Paragraph("Email: espamformationunicampus2@gmail.com", subtext_style))
        elements.append(Spacer(1, 10))

        # Report Title
        report_title_style = ParagraphStyle('ReportTitle', parent=styles['Heading2'], fontSize=14, alignment=1, spaceAfter=10)
        elements.append(Paragraph("ATTENDANCE SUMMARY REPORT", report_title_style))
        elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%d %B, %Y')}", subtext_style))
        elements.append(Spacer(1, 10))

        # Student Details
        student_details = [
            ["Student Name:", student.full_name],
            ["Matric Number:", student.matric_number],
            ["Department:", student.department],
            ["Level:", student.level]
        ]
        student_table = Table(student_details, colWidths=[133, 370])
        student_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        elements.append(student_table)
        elements.append(Spacer(1, 10))

        # Courses Attended
        courses = Course.objects.filter(attendances__student=student).distinct()
        data = [["#", "Course Name", "Course Code", "Total Attendance"]]
        for index, course in enumerate(courses, start=1):
            total_attendance = Attendance.objects.filter(course=course, student=student).count()
            data.append([index, course.course_title, course.course_code, total_attendance])

        # Create and style table
        table = Table(data, colWidths=[40, 200, 120, 120])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))

        elements.append(table)
        doc.build(elements)

        buffer.seek(0)
        return FileResponse(buffer, as_attachment=True, filename=f"attendance_summary_{student.matric_number}.pdf")

    except Exception as e:
        return HttpResponse(f"Error generating PDF: {str(e)}", status=500)